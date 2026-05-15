"""Export v5 Cold Leads to leads.json seed for the CRM repo."""
import json
from openpyxl import load_workbook

SRC = "/Users/abhimanyusajeevan/Library/Application Support/Claude/local-agent-mode-sessions/61ce094c-8b32-422d-8926-06965bb13f15/287e45e4-b1ea-4b1e-af46-46266c74ed16/local_34659fa9-1bed-497a-a6de-d396f2120465/outputs/abhimanyu_cold_leads_500_v5.xlsx"
DST = "/Users/abhimanyusajeevan/Downloads/Claude Team/Code/sponsorship-crm/leads.json"

wb = load_workbook(SRC, data_only=False)
ws = wb["Cold Leads"]

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
# Map original headers to camelCase keys
keymap = {
    "#": "id",
    "Company": "company",
    "Category": "category",
    "HQ Location": "hq",
    "Target Contact Title": "contactTitle",
    "Package Tier": "tier",
    "Draft Code": "draft",
    "Notes / Pitch Angle": "pitchAngle",
    "Outreach Strategy": "channel",
    "Agency / Route": "agency",
    "Contact Pointer": "contactPointer",
    "Priority": "priority",
    "Alignment (1-5)": "alignment",
    "Capacity (1-5)": "capacity",
    "Total Score": "score",
    "Status": "status",
    "Touches Sent (0-9)": "touches",
    "Last Touch Date": "lastTouch",
    "Next Action Date": "nextAction",
    "Response?": "response",
    "Owner": "owner",
    "CRM Notes / Next Step": "notes",
}

leads = []
for r in range(2, ws.max_row + 1):
    row = {}
    for c, h in enumerate(headers, 1):
        key = keymap.get(h, h)
        v = ws.cell(row=r, column=c).value
        if key == "score" and isinstance(v, str) and v.startswith("="):
            # formula → compute from align+cap below
            v = None
        row[key] = v
    # Compute score if missing
    if row.get("score") is None:
        try:
            row["score"] = int(row.get("alignment", 0)) + int(row.get("capacity", 0))
        except Exception:
            row["score"] = 0
    # Normalise None → "" for string cols, 0 for numeric
    for k in ("lastTouch","nextAction","owner","notes"):
        if row.get(k) is None:
            row[k] = ""
    leads.append(row)

with open(DST, "w") as f:
    json.dump({"leads": leads, "exportedAt": "2026-05-15"}, f, indent=2, ensure_ascii=False, default=str)

print(f"Wrote {len(leads)} leads → {DST}")
print("Sample keys:", list(leads[0].keys()))
